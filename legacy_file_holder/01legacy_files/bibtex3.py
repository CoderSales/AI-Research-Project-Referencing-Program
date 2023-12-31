# import numpy
# import pandas
# import openpyxl
from openpyxl import Workbook
import datetime
import os
import shutil
import os.path
wb = Workbook()
ws = wb.active
treeData = [["Type", "Leaf Color", "Height"], ["Maple", "Red", 549], ["Oak", "Green", 783], ["Pine", "Green", 1204]]
for row in treeData:
    ws.append(row)

your_path = os.environ.get("PWD")
root = your_path
files = [(path,f) for path,_,file_list in os.walk(root) for f in file_list]
f_name="Bibtex.bib"
from shutil import  copy
f = open(os.path.join(your_path, f_name), 'r', encoding="utf8")
for i in f.readlines():
    print(i)
f.close()


from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws['A1'] = 42
ws.append([1, 2, 3])
import datetime
ws['A2'] = datetime.datetime.now()
wb.save("sample.xlsx")
